import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 获取资源路径（兼容打包后的exe）
def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


class EmailMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("国科财通steam账号工具 v1.1")
        self.root.geometry("900x600")

        # 设置窗口图标
        try:
            icon_path = get_resource_path('logo.ico')
            self.root.iconbitmap(icon_path)
        except Exception:
            pass  # 图标加载失败时忽略

        # 数据存储 - 辅助邮箱标签页
        self.original_accounts = []  # 存储原始账号 [(email, password), ...]
        self.backup_accounts = []    # 存储辅助邮箱 [(email, password), ...]

        # 数据存储 - Steam账号标签页
        self.steam_accounts = []     # 存储Steam账号 [(username, password), ...]

        # 数据存储 - 标签筛选令牌标签页
        self.token_groups = []      # 存储标签分组 [(file_path, owner_name, account_ids), ...]
        self.token_dir = ""         # 令牌目录

        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        # 创建标签页
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 第一个标签页 - 辅助邮箱
        self.email_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.email_tab, text="辅助邮箱")
        self.setup_email_tab()

        # 第二个标签页 - 导出Steam账号
        self.steam_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.steam_tab, text="导出账号")
        self.setup_steam_tab()

        # 第三个标签页 - 筛选令牌
        self.token_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.token_tab, text="筛选令牌")
        self.setup_token_tab()

    def setup_email_tab(self):
        """设置辅助邮箱标签页"""
        # 顶部按钮区
        button_frame = ttk.Frame(self.email_tab, padding="10")
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="导入原始账号", command=self.import_original).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导入辅助邮箱", command=self.import_backup).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清空数据", command=self.clear_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出合并数据", command=self.export_data).pack(side=tk.LEFT, padx=5)

        # 状态栏
        self.email_status_label = ttk.Label(button_frame, text="等待导入数据...")
        self.email_status_label.pack(side=tk.RIGHT, padx=10)

        # 表格区
        table_frame = ttk.Frame(self.email_tab, padding="10")
        table_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表格
        columns = ("original_email", "original_password", "backup_email", "backup_password")
        self.email_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

        self.email_tree.heading("original_email", text="原始邮箱")
        self.email_tree.heading("original_password", text="原始密码")
        self.email_tree.heading("backup_email", text="辅助邮箱")
        self.email_tree.heading("backup_password", text="辅助密码")

        # 设置列宽
        self.email_tree.column("original_email", width=250)
        self.email_tree.column("original_password", width=150)
        self.email_tree.column("backup_email", width=250)
        self.email_tree.column("backup_password", width=150)

        # 滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.email_tree.yview)
        self.email_tree.configure(yscrollcommand=scrollbar.set)

        self.email_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_steam_tab(self):
        """设置Steam账号标签页"""
        # 顶部按钮区
        button_frame = ttk.Frame(self.steam_tab, padding="10")
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="导入Steam账号", command=self.import_steam).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清空数据", command=self.clear_steam_data).pack(side=tk.LEFT, padx=5)

        # 分隔符选择（下拉框）
        ttk.Label(button_frame, text="分隔符:").pack(side=tk.LEFT, padx=(20, 5))
        self.steam_separator_combo = ttk.Combobox(button_frame, values=["----", ":"], width=8, state="readonly")
        self.steam_separator_combo.set(":")  # 默认值
        self.steam_separator_combo.pack(side=tk.LEFT, padx=5)

        ttk.Button(button_frame, text="导出Steam账号", command=self.export_steam).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出xlsx", command=self.export_steam_xlsx).pack(side=tk.LEFT, padx=5)

        # 状态栏
        self.steam_status_label = ttk.Label(button_frame, text="等待导入数据...")
        self.steam_status_label.pack(side=tk.RIGHT, padx=10)

        # 表格区
        table_frame = ttk.Frame(self.steam_tab, padding="10")
        table_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表格
        columns = ("username", "password")
        self.steam_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

        self.steam_tree.heading("username", text="账号")
        self.steam_tree.heading("password", text="密码")

        # 设置列宽
        self.steam_tree.column("username", width=400)
        self.steam_tree.column("password", width=400)

        # 滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.steam_tree.yview)
        self.steam_tree.configure(yscrollcommand=scrollbar.set)

        self.steam_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_token_tab(self):
        """设置标签筛选令牌标签页"""
        # 顶部说明
        info_label = ttk.Label(
            self.token_tab,
            text="通过账号文件筛选令牌：将指定账号的令牌复制到对应归属的目录",
            padding="10"
        )
        info_label.pack(fill=tk.X)

        # 账号文件和归属输入区
        self.token_input_frame = ttk.LabelFrame(self.token_tab, text="账号文件与归属", padding="10")
        self.token_input_frame.pack(fill=tk.X, padx=10, pady=5)

        # 账号文件列表显示
        columns = ("file_path", "owner_name", "account_count")
        self.token_file_tree = ttk.Treeview(self.token_input_frame, columns=columns, show="headings", height=6)

        self.token_file_tree.heading("file_path", text="账号文件")
        self.token_file_tree.heading("owner_name", text="归属")
        self.token_file_tree.heading("account_count", text="账号数量")

        self.token_file_tree.column("file_path", width=350)
        self.token_file_tree.column("owner_name", width=100)
        self.token_file_tree.column("account_count", width=80)

        scrollbar = ttk.Scrollbar(self.token_input_frame, orient=tk.VERTICAL, command=self.token_file_tree.yview)
        self.token_file_tree.configure(yscrollcommand=scrollbar.set)

        self.token_file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 添加账号文件按钮区
        add_frame = ttk.Frame(self.token_input_frame)
        add_frame.pack(fill=tk.X, pady=5)

        ttk.Button(add_frame, text="添加账号文件", command=self.add_token_account_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(add_frame, text="删除选中", command=self.remove_token_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(add_frame, text="清空全部", command=self.clear_token_files).pack(side=tk.LEFT, padx=5)

        # 令牌目录选择区
        token_dir_frame = ttk.LabelFrame(self.token_tab, text="令牌目录", padding="10")
        token_dir_frame.pack(fill=tk.X, padx=10, pady=5)

        self.token_dir_entry = ttk.Entry(token_dir_frame)
        self.token_dir_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Button(token_dir_frame, text="选择目录", command=self.select_token_dir).pack(side=tk.LEFT, padx=5)

        # 输出目录设置
        output_frame = ttk.LabelFrame(self.token_tab, text="输出目录（各归属的令牌将保存到此目录）", padding="10")
        output_frame.pack(fill=tk.X, padx=10, pady=5)

        self.output_dir_entry = ttk.Entry(output_frame)
        self.output_dir_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Button(output_frame, text="选择目录", command=self.select_output_dir).pack(side=tk.LEFT, padx=5)

        # 状态栏
        self.token_status_label = ttk.Label(self.token_tab, text="请添加账号文件和归属")
        self.token_status_label.pack(pady=5)

        # 执行按钮
        ttk.Button(self.token_tab, text="开始筛选令牌", command=self.filter_tokens).pack(pady=10)

    def add_token_account_file(self):
        """添加账号文件和归属"""
        file_path = filedialog.askopenfilename(
            title="选择账号文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        # 弹出对话框让用户输入归属
        dialog = tk.Toplevel(self.root)
        dialog.title("输入归属")
        dialog.geometry("300x120")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="请输入此账号文件的归属名称：").pack(pady=10)

        owner_entry = ttk.Entry(dialog, width=30)
        owner_entry.pack(pady=10)
        owner_entry.focus()

        result = {"owner": None}

        def on_ok():
            result["owner"] = owner_entry.get().strip()
            if result["owner"]:
                dialog.destroy()
            else:
                messagebox.showwarning("提示", "请输入归属名称！")

        def on_cancel():
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT, padx=5)

        self.root.wait_window(dialog)

        if not result["owner"]:
            return

        # 读取账号文件，获取账号ID列表
        try:
            account_ids = set()
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # 支持两种格式：账号ID----密码 或 账号ID:密码
                    if "----" in line:
                        parts = line.split("----")
                    else:
                        parts = line.split(":")

                    if parts:
                        account_id = parts[0].strip()
                        if account_id:
                            account_ids.add(account_id)

            # 保存到列表
            self.token_groups.append({
                "file_path": file_path,
                "owner": result["owner"],
                "account_ids": account_ids
            })

            self.update_token_file_tree()
            self.token_status_label.config(
                text=f"已添加 {result['owner']}，包含 {len(account_ids)} 个账号"
            )

        except Exception as e:
            messagebox.showerror("错误", f"读取账号文件失败：{str(e)}")

    def update_token_file_tree(self):
        """更新账号文件列表显示"""
        for item in self.token_file_tree.get_children():
            self.token_file_tree.delete(item)

        for group in self.token_groups:
            self.token_file_tree.insert("", tk.END, values=(
                os.path.basename(group["file_path"]),
                group["owner"],
                len(group["account_ids"])
            ))

    def remove_token_file(self):
        """删除选中的账号文件"""
        selected = self.token_file_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选中要删除的项！")
            return

        idx = self.token_file_tree.index(selected[0])
        self.token_groups.pop(idx)
        self.update_token_file_tree()
        self.token_status_label.config(text="已删除选中的账号文件")

    def clear_token_files(self):
        """清空所有账号文件"""
        if messagebox.askyesno("确认", "确定要清空所有账号文件吗？"):
            self.token_groups = []
            self.update_token_file_tree()
            self.token_status_label.config(text="已清空所有账号文件")

    def select_token_dir(self):
        """选择令牌目录"""
        dir_path = filedialog.askdirectory(title="选择令牌目录")
        if dir_path:
            self.token_dir = dir_path
            self.token_dir_entry.delete(0, tk.END)
            self.token_dir_entry.insert(0, dir_path)

    def select_output_dir(self):
        """选择输出目录"""
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.output_dir_entry.delete(0, tk.END)
            self.output_dir_entry.insert(0, dir_path)

    def filter_tokens(self):
        """筛选令牌"""
        if not self.token_groups:
            messagebox.showwarning("提示", "请先添加账号文件和归属！")
            return

        token_dir = self.token_dir_entry.get().strip()
        if not token_dir or not os.path.isdir(token_dir):
            messagebox.showwarning("提示", "请选择有效的令牌目录！")
            return

        output_dir = self.output_dir_entry.get().strip()
        if not output_dir:
            messagebox.showwarning("提示", "请选择输出目录！")
            return

        # 确保输出目录存在
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                messagebox.showerror("错误", f"创建输出目录失败：{str(e)}")
                return

        # 获取令牌目录中的所有.maFile文件
        mafiles = {}
        try:
            for filename in os.listdir(token_dir):
                if filename.endswith(".maFile"):
                    # 从文件名提取账号ID
                    account_id = filename.replace(".maFile", "")
                    mafiles[account_id] = os.path.join(token_dir, filename)
        except Exception as e:
            messagebox.showerror("错误", f"读取令牌目录失败：{str(e)}")
            return

        # 为每个归属创建目录并复制令牌
        results = []
        total_copied = 0

        for group in self.token_groups:
            owner = group["owner"]
            account_ids = group["account_ids"]

            # 创建归属目录
            owner_dir = os.path.join(output_dir, f"{owner}令牌")
            if not os.path.exists(owner_dir):
                os.makedirs(owner_dir)

            # 匹配并复制令牌
            copied_count = 0
            not_found = []

            for account_id in account_ids:
                if account_id in mafiles:
                    src = mafiles[account_id]
                    dst = os.path.join(owner_dir, f"{account_id}.maFile")
                    try:
                        import shutil
                        shutil.copy2(src, dst)
                        copied_count += 1
                    except Exception as e:
                        pass
                else:
                    not_found.append(account_id)

            results.append(f"{owner}: 找到 {copied_count} 个令牌")
            total_copied += copied_count

        # 显示结果
        result_text = "\n".join(results)
        self.token_status_label.config(text=f"完成！共复制 {total_copied} 个令牌")
        messagebox.showinfo("完成", f"令牌筛选完成！\n\n{result_text}")

    def import_original(self):
        """导入原始账号文件"""
        file_path = filedialog.askopenfilename(
            title="选择原始账号文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            self.original_accounts = []
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # 使用 ---- 分割
                    parts = line.split("----")
                    if len(parts) >= 4:
                        email = parts[2]  # 第三个字段是邮箱
                        password = parts[3]  # 第四个字段是密码
                        self.original_accounts.append((email, password))

            self.update_email_table()
            self.email_status_label.config(text=f"已导入 {len(self.original_accounts)} 个原始账号")

        except Exception as e:
            messagebox.showerror("错误", f"导入原始账号失败：{str(e)}")

    def import_backup(self):
        """导入辅助邮箱文件"""
        if not self.original_accounts:
            messagebox.showwarning("提示", "请先导入原始账号！")
            return

        file_path = filedialog.askopenfilename(
            title="选择辅助邮箱文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            self.backup_accounts = []
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # 使用 --- 分割
                    parts = line.split("---")
                    if len(parts) >= 2:
                        email = parts[0]  # 第一个字段是邮箱
                        password = parts[1]  # 第二个字段是密码
                        self.backup_accounts.append((email, password))

            # 一对一匹配
            matched_count = min(len(self.original_accounts), len(self.backup_accounts))

            self.update_email_table()
            self.email_status_label.config(
                text=f"已导入 {len(self.backup_accounts)} 个辅助邮箱，成功匹配 {matched_count} 个"
            )

            if len(self.backup_accounts) < len(self.original_accounts):
                messagebox.showinfo("提示",
                    f"辅助邮箱数量({len(self.backup_accounts)})少于原始账号数量({len(self.original_accounts)})，\n"
                    f"只有前 {matched_count} 个账号被匹配。")

        except Exception as e:
            messagebox.showerror("错误", f"导入辅助邮箱失败：{str(e)}")

    def _clear_tree(self, tree):
        """清空Treeview表格"""
        for item in tree.get_children():
            tree.delete(item)

    def _check_export_data(self, data, status_label):
        """检查是否有数据可导出"""
        if not data:
            messagebox.showwarning("提示", "没有数据可导出！")
            return False
        return True

    def update_email_table(self):
        """更新邮箱标签页表格显示"""
        self._clear_tree(self.email_tree)

        max_rows = max(len(self.original_accounts), len(self.backup_accounts))

        for i in range(max_rows):
            orig = self.original_accounts[i] if i < len(self.original_accounts) else ("", "")
            backup = self.backup_accounts[i] if i < len(self.backup_accounts) else ("", "")

            self.email_tree.insert("", tk.END, values=(
                orig[0], orig[1], backup[0], backup[1]
            ))

    def clear_data(self):
        """清空邮箱标签页数据"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            self.original_accounts = []
            self.backup_accounts = []
            self.update_email_table()
            self.email_status_label.config(text="数据已清空")

    def export_data(self):
        """导出合并数据"""
        if not self.original_accounts or not self.backup_accounts:
            messagebox.showwarning("提示", "没有数据可导出！")
            return

        file_path = filedialog.asksaveasfilename(
            title="保存合并数据",
            initialfile="邮箱辅助关系.txt",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                max_rows = min(len(self.original_accounts), len(self.backup_accounts))

                for i in range(max_rows):
                    orig_email, orig_password = self.original_accounts[i]
                    backup_email, backup_password = self.backup_accounts[i]

                    # 导出格式：原始邮箱----原始邮箱密码----辅助邮箱----辅助邮箱密码
                    f.write(f"{orig_email}----{orig_password}----{backup_email}----{backup_password}\n")

            self.email_status_label.config(text=f"已导出 {max_rows} 条记录到 {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"成功导出 {max_rows} 条记录！")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def import_steam(self):
        """导入Steam账号"""
        file_path = filedialog.askopenfilename(
            title="选择原始账号文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            self.steam_accounts = []
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # 使用 ---- 分割
                    parts = line.split("----")
                    if len(parts) >= 2:
                        username = parts[0]  # 第一个字段是账号
                        password = parts[1]  # 第二个字段是密码
                        self.steam_accounts.append((username, password))

            self.update_steam_table()
            self.steam_status_label.config(text=f"已导入 {len(self.steam_accounts)} 个Steam账号")

        except Exception as e:
            messagebox.showerror("错误", f"导入Steam账号失败：{str(e)}")

    def update_steam_table(self):
        """更新Steam账号表格显示"""
        self._clear_tree(self.steam_tree)

        for username, password in self.steam_accounts:
            self.steam_tree.insert("", tk.END, values=(username, password))

    def clear_steam_data(self):
        """清空Steam账号数据"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            self.steam_accounts = []
            self.update_steam_table()
            self.steam_status_label.config(text="数据已清空")

    def export_steam(self):
        """导出Steam账号"""
        if not self._check_export_data(self.steam_accounts, self.steam_status_label):
            return

        file_path = filedialog.asksaveasfilename(
            title="保存Steam账号",
            initialfile="steam账号.txt",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            separator = self.steam_separator_combo.get()
            with open(file_path, "w", encoding="utf-8") as f:
                for username, password in self.steam_accounts:
                    f.write(f"{username}{separator}{password}\n")

            self.steam_status_label.config(text=f"已导出 {len(self.steam_accounts)} 条记录到 {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"成功导出 {len(self.steam_accounts)} 条记录！")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def export_steam_xlsx(self):
        """导出Steam账号到xlsx文件"""
        if not self._check_export_data(self.steam_accounts, self.steam_status_label):
            return

        if not HAS_OPENPYXL:
            messagebox.showerror("错误", "需要安装openpyxl库才能导出xlsx文件！\n请运行: pip install openpyxl")
            return

        file_path = filedialog.asksaveasfilename(
            title="保存Steam账号",
            initialfile="账号密码.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "账号密码"

            # 写入表头
            ws["A1"] = "账号"
            ws["B1"] = "密码"

            # 写入数据
            for idx, (username, password) in enumerate(self.steam_accounts, start=2):
                ws[f"A{idx}"] = username
                ws[f"B{idx}"] = password

            # 调整列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30

            wb.save(file_path)

            self.steam_status_label.config(text=f"已导出 {len(self.steam_accounts)} 条记录到 {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"成功导出 {len(self.steam_accounts)} 条记录到xlsx文件！")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")


def main():
    root = tk.Tk()
    app = EmailMergerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
